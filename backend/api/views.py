from rest_framework import generics, status, serializers, viewsets
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.auth import  update_session_auth_hash
from .models import Enseignant,Etudiant, HelpRequest,Salle, Presence, QRNotification, Session, Filiere, Niveau, Matiere
from .serializers import EnseignantSignupSerializer, CustomTokenObtainPairSerializer, EtudiantRegisterSerializer, SessionSerializer, PasswordSerializer, EmailSerializer, HelpRequestSerializer, PresenceSerializer, FiliereSerializer, MatiereSerializer, NiveauSerializer, EnseignantSerializer, EnseignantCreateUpdateSerializer, EnseignantListSerializer, SalleSerializer
from .serializers import EtudiantSerializer
import qrcode
import re
from django.utils.dateparse import parse_date
import csv
from django.core.files import File
from openpyxl.chart import BarChart, Reference
from io import BytesIO
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.exceptions import AuthenticationFailed
from rest_framework.authtoken.models import Token
from rest_framework.decorators import api_view, action
from datetime import datetime, timedelta
from rest_framework import permissions
from django.core.mail import send_mail
from rest_framework.generics import get_object_or_404
import pytz
from django.http import HttpResponse
import pandas as pd
from django.utils.timezone import now as timezone_now
from openpyxl import Workbook
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.tokens import RefreshToken
from dateutil.relativedelta import relativedelta
from django.db.models import Q, Count
from rest_framework import filters
from openpyxl.utils.dataframe import dataframe_to_rows
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.utils import timezone
from django.utils.timezone import now
from openpyxl.styles import Font, Alignment
import calendar
from langchain_fireworks import ChatFireworks
from collections import Counter, defaultdict

###############################L'authentification################################################


#####################Signup  d'enseignant#######################################

class EnseignantSignupView(generics.CreateAPIView):
    queryset = Enseignant.objects.all()
    serializer_class = EnseignantSignupSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            user = serializer.save()
            enseignant = user.enseignant
           
            
            return Response(
                {
                    "message": "Compte créé avec succès . Vous pouvez maintenant vous connecter.",
                    "statut": enseignant.statut
                },
                status=status.HTTP_201_CREATED
            )
        except serializers.ValidationError as e:
            print("Serializer errors:", serializer.errors)
            return Response(
                {"error": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )
##################Signup d'etudiant############################

class RegisterEtudiantView(APIView):
    permission_classes = [AllowAny]
    def post(self, request):
        serializer = EtudiantRegisterSerializer(data=request.data)
        if serializer.is_valid():
            etudiant = serializer.save()
            return Response(
                {"message": "Inscription réussie"}, 
                status=status.HTTP_201_CREATED
            )
        return Response(
            {"error": "Données invalides", "details": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )

##########Login
class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)

        try:
            serializer.is_valid(raise_exception=True)
        except Exception as e:
            return Response({'detail': 'Email ou mot de passe incorrect'}, status=status.HTTP_401_UNAUTHORIZED)
    

        return Response(serializer.validated_data, status=status.HTTP_200_OK)

######################GENERATION DL CODE W SCAN#########################################################

#################Genenation qr code

class GenerateQRCodeAPI(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not hasattr(request.user, 'enseignant'):
            return Response({"error": "Utilisateur non lié à un enseignant."}, status=403)

        user = request.user
        enseignant = user.enseignant
        if enseignant.statut.lower() not in ['professeur', 'vacataire']:
            return Response({"error": "Seuls les enseignants peuvent générer des QR codes."}, status=403)

        session_id = request.data.get('session_id')
        if session_id:
            try:
                session = Session.objects.get(id=session_id, enseignant=user)
            except Session.DoesNotExist:
                return Response({"error": "Séance introuvable."}, status=404)
        else:
            now = timezone.now()
            today = now.date()
            current_time = now.replace(second=0, microsecond=0).time()
            jour = now.strftime('%A')
            jour_mapping = {
                'Monday': 'Lundi', 'Tuesday': 'Mardi', 'Wednesday': 'Mercredi',
                'Thursday': 'Jeudi', 'Friday': 'Vendredi', 'Saturday': 'Samedi', 'Sunday': 'Dimanche'
            }
            jour = jour_mapping.get(jour, jour)

            session = Session.objects.filter(
                enseignant=user,
                jour=jour,
                date_debut__lte=today,
                date_fin__gte=today,
                heure_debut__lte=current_time,
                heure_fin__gte=current_time
            ).first()

            if not session:
                return Response({"error": "Aucune séance en cours trouvée."}, status=404)

        # Vérifier si un QR code existe déjà et est encore valide
        if session.qr_code and session.qr_generated_at:
            if timezone.now() - session.qr_generated_at < timedelta(minutes=10):
                return Response({
                    "success": True,
                    "qr_code_url": request.build_absolute_uri(session.qr_code.url),
                    "session_id": session.id,
                    "message": "QR code déjà généré récemment."
                })

        # Générer un nouveau QR code
        qr_data = str(session.id)
        qr_img = qrcode.make(qr_data)
        buffer = BytesIO()
        qr_img.save(buffer, format="PNG")
        session.qr_code.save(f"qr_{session.id}.png", File(buffer))
        session.qr_generated_at = timezone.now()
        session.save()

        # Créer une notification li katmchi l admin bach y3ref seance o taux de presence
        QRNotification.objects.create(
            enseignant=enseignant,
            session=session,
            viewed=False
        )

        return Response({
            "success": True,
            "qr_code_url": request.build_absolute_uri(session.qr_code.url),
            "session_id": session.id
        })
        
######Recuperation de la liste des etudiants
        
@api_view(['GET'])
def get_absences(request):
    user = request.user
    if not hasattr(user, 'enseignant'):
        return Response({"error": "Accès réservé aux enseignants"}, status=403)

    session_id = request.query_params.get('session_id')
    if session_id:
        try:
            session = Session.objects.get(id=session_id, enseignant=user)
        except Session.DoesNotExist:
            return Response({"error": "Séance introuvable."}, status=404)
    else:
        now = timezone.now()
        today = now.date()
        current_time = now.replace(second=0, microsecond=0).time()
        jour = now.strftime('%A')
        jour_mapping = {
            'Monday': 'Lundi', 'Tuesday': 'Mardi', 'Wednesday': 'Mercredi',
            'Thursday': 'Jeudi', 'Friday': 'Vendredi', 'Saturday': 'Samedi', 'Sunday': 'Dimanche'
        }
        jour = jour_mapping.get(jour, jour)

        session = Session.objects.filter(
            enseignant=user,
            jour=jour,
            date_debut__lte=today,
            date_fin__gte=today,
            heure_debut__lte=current_time,
            heure_fin__gte=current_time
        ).first()

        if not session:
            return Response({"error": "Aucune séance en cours trouvée."}, status=404)

    filiere = session.matiere.filiere
    niveau = session.matiere.niveau

    etudiants = Etudiant.objects.filter(filiere=filiere, niveau=niveau).order_by('nom', 'prenom')
    presences = Presence.objects.filter(session=session)
    presence_dict = {p.etudiant_id: p for p in presences}

    result = []
    for etudiant in etudiants:
        presence = presence_dict.get(etudiant.id)
        result.append({
            "nom": etudiant.nom,
            "prenom": etudiant.prenom,
            "status": "présent(e)" if presence else "absent(e)",
            "justifiee": presence.justifiee if presence else False,
            "scanned_at": presence.scanned_at if presence else None
        })

    total = len(result)
    presents = sum(1 for r in result if r["status"] == "présent(e)")
    absents = total - presents
    taux = round((presents / total) * 100, 2) if total > 0 else 0

    return Response({
        "session_id": session.id,
        "filiere": filiere.nom,
        "niveau": niveau.nom,
        "etudiants": result,
        "statistiques": {
            "presences": presents,
            "absences": absents,
            "taux": taux
        }
    })

#######################Hadi la vue dyalk d enregistrer la presence li mt3l9a bscan modifit ghi chi dfchat 3la 7sab simple jwt

@api_view(['POST'])
def enregistrer_presence_scan(request):
    user = request.user

    if user.role != 'etudiant':
        return Response({"error": "Accès non autorisé"}, status=403)

    session_id = request.data.get('session_id')
    if not session_id:
        return Response({"error": "ID de session manquant"}, status=400)

    try:
        session = Session.objects.get(id=session_id)
    except Session.DoesNotExist:
        return Response({"error": "Session introuvable"}, status=404)

    
    etudiant = user.etudiant

    presence, created = Presence.objects.get_or_create(
        etudiant=etudiant,
        session=session,
        defaults={'status': 'présent(e)'}
    )

    if created:
        return Response({
            "message": "Présence enregistrée",
            "etudiant": f"{etudiant.prenom} {etudiant.nom}",
            "session": session.id,
            "scanned_at": presence.scanned_at
        }, status=201)
    else:
        return Response({
            "message": "Présence déjà enregistrée",
            "presence_id": presence.id
        }, status=200)

#########################Hadi la vue lokhra dyal li katrecupere biha lpresence ta hiya nfs l7aja 

@api_view(['GET'])
def get_student_presences(request):
    user = request.user
    


    if user.role != 'etudiant':
        return Response({"error": "Accès non autorisé"}, status=403)

    etudiant = user.etudiant
    presences = Presence.objects.filter(etudiant=etudiant).select_related(
        'session__matiere', 'session__salle', 'session__enseignant'
    ).order_by('-session__date_debut')

    result = []
    for presence in presences:
        session = presence.session
        enseignant_obj = Enseignant.objects.get(user=session.enseignant)
        result.append({
            'id': presence.id,
            'matiere': session.matiere.nom,
            'date': session.date_debut,
            'heure_debut': session.heure_debut.strftime('%H:%M'),
            'heure_fin': session.heure_fin.strftime('%H:%M'),
            'status': presence.status,
            'justifiee': presence.justifiee,
            'scanned_at': presence.scanned_at,
            'prof_nom': f"{enseignant_obj.prenom} {enseignant_obj.nom}",
            'salle_nom': session.salle.nom if session.salle else 'Inconnue',
            'type_seance': session.type_seance
        })
    return Response(result, status=200)

#############Had l modele li nayd 3lih lfilm lenseignant y9dr y3mr les seances dyalo mra we7da pdt wa7ed lperiode
############System mn b3d kayb9a ydetecter automatiquement seance: niv fil matiere..............
class SessionViewSet(viewsets.ModelViewSet):
    serializer_class = SessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Session.objects.filter(enseignant=self.request.user)

    def perform_create(self, serializer):
        data = serializer.validated_data
        date_debut = data['date_debut']
        date_fin = data['date_fin']
        jour_cible = data['jour']

        jours_map = {
            'Lundi': 0, 'Mardi': 1, 'Mercredi': 2,
            'Jeudi': 3, 'Vendredi': 4, 'Samedi': 5
        }
        weekday_target = jours_map.get(jour_cible)

        current = date_debut
        while current <= date_fin:
            if current.weekday() == weekday_target:
                Session.objects.create(
                    enseignant=self.request.user,
                    matiere=data['matiere'],
                    salle=data['salle'],
                    jour=jour_cible,
                    heure_debut=data['heure_debut'],
                    heure_fin=data['heure_fin'],
                    date_debut=current,
                    date_fin=current
                )
            current += timedelta(days=1)

############Les vues d lenseignant##############################################
### Katregrouper nb detudiants f seances , nbr de seances dlprof o la seance prochaine #####################



class DashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        now = timezone_now().astimezone(pytz.timezone('Africa/Casablanca'))

        today_date = now.date()
        current_time = now.time()

        # Récupérer toutes les séances à venir (y compris aujourd'hui)
        sessions = Session.objects.filter(
            enseignant=user,
            date_debut__gte=today_date
        ).order_by('date_debut', 'heure_debut')

        selected_session = None
        for session in sessions:
            session_datetime = datetime.combine(session.date_debut, session.heure_debut).astimezone(pytz.timezone('Africa/Casablanca'))
            if session.date_debut == today_date:
                if session.heure_debut <= current_time <= session.heure_fin:
                    selected_session = session
                    break
                elif current_time < session.heure_debut:
                    selected_session = session
                    break
            elif session.date_debut > today_date:
                selected_session = session
                break

        # Initialiser les compteurs
        nb_etudiants = 0
        nb_seances = 0
        session_data = None

        if selected_session:
            filiere = selected_session.matiere.filiere
            niveau = selected_session.matiere.niveau

            nb_etudiants = Etudiant.objects.filter(
                filiere=filiere,
                niveau=niveau
            ).count()

            nb_seances = Session.objects.filter(
                enseignant=user,
                matiere__filiere=filiere,
                matiere__niveau=niveau
            ).count()

            session_data = {
                'id': selected_session.id,
                'module': selected_session.matiere.nom,
                'heure_debut': selected_session.heure_debut.strftime('%H:%M'),
                'heure_fin': selected_session.heure_fin.strftime('%H:%M'),
                'jour': selected_session.jour,
                'date': selected_session.date_debut.strftime('%Y-%m-%d'),
                'salle': selected_session.salle.nom,
                'filiere': filiere.nom,
                'niveau': niveau.nom,
            }
            try:
                enseignant = user.enseignant
                nom = enseignant.nom
                prenom = enseignant.prenom
            except:
                nom = user.username or ''
                prenom = ''


        return Response({
            'nom': nom,
            'prenom': prenom,
            'nb_etudiants': nb_etudiants,
            'nb_seances': nb_seances,
            'seance_prochaine': session_data,
        })


#############Hada ghi 3la 7sab l modification dl profile: email w password
###########ladmin li kaykn dayr password mowe7ed mais mn b3d lenseignant kaybdlo
            
class ProfileAPI(APIView):
    permission_classes = [IsAuthenticated]

    # Get current email
    def get(self, request):
        return Response({'email': request.user.email})

    # Update email
    def post(self, request):
        serializer = EmailSerializer(data=request.data)
        if serializer.is_valid():
            request.user.email = serializer.validated_data['email']
            request.user.save()
            return Response({'status': 'Email updated'})
        return Response(serializer.errors, status=400)

class PasswordAPI(APIView):
    permission_classes = [IsAuthenticated]

    # Update password
    def post(self, request):
        serializer = PasswordSerializer(data=request.data)
        if serializer.is_valid():
            if not request.user.check_password(serializer.validated_data['old_password']):
                return Response({'error': 'Wrong current password'}, status=400)
            
            request.user.set_password(serializer.validated_data['new_password'])
            request.user.save()
            update_session_auth_hash(request, request.user)  # Prevent logout
            return Response({'status': 'Password updated'})
        return Response(serializer.errors, status=400)

# rapport cote enseignant 
#  Nettoyage du texte IA
def nettoyer_texte_ia(texte):
    return re.sub(r"<think>.*?</think>", "", texte, flags=re.DOTALL).strip()

# Générer un résumé IA pour l’enseignant
def generer_resume_enseignant(absences_data, enseignant):
    if not absences_data:
        return "Aucune absence à analyser pour cette période."

    resume = "\n".join([
        f"- {a['Nom']} {a['Prénom']} | {a['Module']} | {a['Date']} | Justifiée : {a['Justifiée']}"
        for a in absences_data
    ])

    prompt = f"""
Voici les données d'absences des étudiants du module  :

{resume}

Analyse uniquement ces données. Fournis un résumé structuré avec :
1. Les étudiants les plus absents
2. Les séances ou dates critiques
3. Des recommandations pédagogiques
4. Une conclusion synthétique

⚠️ Utilise uniquement les noms et dates présents dans les données. Ne crée pas d’exemples fictifs.

Réponds en français.
"""
    try:
        llm = ChatFireworks(model="accounts/fireworks/models/deepseek-r1", temperature=0.3)
        response = llm.invoke(prompt)
        return nettoyer_texte_ia(response.content)
    except:
        return "Résumé IA indisponible pour le moment."

class ExportRapportAbsencesEnseignant(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        filiere_id = request.query_params.get('filiere_id')
        niveau_id = request.query_params.get('niveau_id')
        matiere_id = request.query_params.get('matiere_id')
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')

        if not all([filiere_id, niveau_id, matiere_id, date_debut, date_fin]):
            return Response({"error": "Tous les paramètres sont requis."}, status=400)

        try:
            date_debut = datetime.strptime(date_debut, "%Y-%m-%d").date()
            date_fin = datetime.strptime(date_fin, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Format de date invalide. Utilisez YYYY-MM-DD."}, status=400)

        sessions = Session.objects.filter(
            enseignant=user,
            matiere_id=matiere_id,
            date_debut__gte=date_debut,
            date_fin__lte=date_fin
        ).select_related('matiere', 'salle')

        if not sessions.exists():
            return Response({"error": "Aucune séance trouvée pour cette période."}, status=404)

        matiere = sessions.first().matiere
        filiere = matiere.filiere
        niveau = matiere.niveau

        etudiants = Etudiant.objects.filter(
            filiere=filiere,
            niveau=niveau
        ).select_related('user').order_by('nom', 'prenom')

        absences_data = []
        for session in sessions:
            presences = Presence.objects.filter(session=session)
            presents_ids = set(p.etudiant_id for p in presences)

            for etudiant in etudiants:
                if etudiant.id not in presents_ids:
                    absences_data.append({
                        'Nom': etudiant.nom,
                        'Prénom': etudiant.prenom,
                        'Module': session.matiere.nom,
                        'Type': session.type_seance,
                        'Date': session.date_debut.strftime('%d/%m/%Y'),
                        'Heure': f"{session.heure_debut.strftime('%H:%M')} - {session.heure_fin.strftime('%H:%M')}",
                        'Salle': session.salle.nom,
                        'Justifiée': 'Non'
                    })

        # Grouper les absences par étudiant
        absences_par_etudiant = defaultdict(list)
        for absence in absences_data:
            key = (absence['Nom'], absence['Prénom'])
            absences_par_etudiant[key].append(absence)

        wb = Workbook()
        wb.remove(wb.active)

        # Feuille Étudiants
        ws_etudiants = wb.create_sheet("Étudiants")
        ws_etudiants.append(["Nom", "Prénom", "Email"])
        for e in etudiants:
            email = e.user.email if hasattr(e, 'user') and e.user else ''
            ws_etudiants.append([e.nom, e.prenom, email])

        # Feuille Absences
        ws_absences = wb.create_sheet("Absences")
        if absences_data:
            ws_absences.append(list(absences_data[0].keys()))
            for row in absences_data:
                ws_absences.append(list(row.values()))
        else:
            ws_absences.append(["Aucune absence enregistrée."])

        # Feuille Statistiques
        ws_stats = wb.create_sheet("Statistiques")
        if absences_data:
            df = pd.DataFrame(absences_data)
            stats = df.groupby('Module').size().reset_index(name='Total Absences')
            stats['Taux Absence (%)'] = (stats['Total Absences'] / len(etudiants) * 100).round(2)

            ws_stats.append(["Module", "Total Absences", "Taux Absence (%)"])
            for _, row in stats.iterrows():
                ws_stats.append([row['Module'], row['Total Absences'], row['Taux Absence (%)']])

            chart = BarChart()
            chart.title = "Absences par module"
            chart.x_axis.title = "Module"
            chart.y_axis.title = "Total Absences"
            data = Reference(ws_stats, min_col=2, max_col=2, min_row=1, max_row=ws_stats.max_row)
            categories = Reference(ws_stats, min_col=1, min_row=2, max_row=ws_stats.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            ws_stats.add_chart(chart, "E2")
        else:
            ws_stats.append(["Aucune donnée statistique disponible."])

        # Résumé IA
        texte_ia = generer_resume_enseignant(absences_data, user)
        ws_ia = wb.create_sheet("Résumé IA")
        ws_ia["A1"] = "Analyse IA personnalisée"
        ws_ia["A1"].font = Font(bold=True, size=14)
        for i, ligne in enumerate(texte_ia.splitlines(), start=2):
            ws_ia[f"A{i}"] = ligne
            ws_ia[f"A{i}"].alignment = Alignment(wrap_text=True)

        #  Étudiants à risque
        compteur = Counter((a['Nom'], a['Prénom']) for a in absences_data if a['Justifiée'] == 'Non')
        ws_risque = wb.create_sheet("Étudiants à risque")
        ws_risque.append(["Nom", "Prénom", "Absences non justifiées"])
        for (nom, prenom), total in compteur.items():
            if total >= 3:
                ws_risque.append([nom, prenom, total])

        #  Feuilles individuelles pour chaque étudiant
        for (nom, prenom), absences in absences_par_etudiant.items():
            feuille_nom = f"{nom}_{prenom}"[:31]  # max 31 caractères
            ws_indiv = wb.create_sheet(title=feuille_nom)
            ws_indiv.append(["Module", "Type", "Date", "Heure", "Salle", "Justifiée"])
            for a in absences:
                ws_indiv.append([a['Module'], a['Type'], a['Date'], a['Heure'], a['Salle'], a['Justifiée']])
            ws_indiv.append([])
            ws_indiv.append(["Résumé :"])
            ws_indiv.append([f"Total absences : {len(absences)}"])
            non_just = sum(1 for a in absences if a['Justifiée'] == 'Non')
            ws_indiv.append([f"Absences non justifiées : {non_just}"])

        # Export
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"rapport_absences_{filiere.nom}_{niveau.nom}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
##########Had l api katjme3 etudiant w enseignant kaysifto biha demande l admin 

class HelpRequestAPI(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if hasattr(request.user, 'enseignant'):
            help_requests = HelpRequest.objects.filter(enseignant=request.user)
        elif hasattr(request.user, 'etudiant'):
            help_requests = HelpRequest.objects.filter(etudiant=request.user)
        else:
            return Response({'error': 'Utilisateur non reconnu.'}, status=400)

        help_requests = help_requests.order_by('-created_at')
        serializer = HelpRequestSerializer(help_requests, many=True)
        return Response(serializer.data)

    def post(self, request):
        subject = request.data.get('subject')
        message = request.data.get('message')

        if not subject or not message:
            return Response({'error': 'Sujet et message requis.'}, status=400)

        help_request = HelpRequest.objects.create(
            subject=subject,
            message=message,
            enseignant=request.user if hasattr(request.user, 'enseignant') else None,
            etudiant=request.user if hasattr(request.user, 'etudiant') else None,
        )
        return Response({'status': 'Message envoyé avec succès !'}, status=201)
#################3Hadi la vue dyalk dl profile wrah tahiya modifit chi 7wayj bach tkhdm 3la 7sab simplejwt oraha khdama 

@api_view(['GET', 'PUT'])

def student_profile_view(request):
    user = request.user

    if user.role != 'etudiant':
        return Response({"error": "Accès non autorisé"}, status=403)

    etudiant = user.etudiant

    if request.method == 'GET':
        return Response({
            'id': etudiant.id,
            'nom': etudiant.nom,
            'prenom': etudiant.prenom,
            'email': user.email,
            'filiere': etudiant.filiere,
            'niveau': etudiant.niveau,
            'photo': etudiant.photo.url if etudiant.photo else None
        })

    elif request.method == 'PUT':
        data = request.data

        # Mise à jour des champs Etudiant
        etudiant.nom = data.get('nom', etudiant.nom)
        etudiant.prenom = data.get('prenom', etudiant.prenom)
        etudiant.filiere = data.get('filiere', etudiant.filiere)
        etudiant.niveau = data.get('niveau', etudiant.niveau)

        # Mise à jour de l'email
        new_email = data.get('email')
        if new_email and new_email != user.email:
            if not re.match(r'^[a-zA-Z0-9_.+-]+@ump\.ac\.ma$', new_email):
                return Response({"error": "Email UMP requis (ex: nom@ump.ac.ma)"}, status=400)
            user.email = new_email

        # Mise à jour du mot de passe
        new_password = data.get('password')
        if new_password:
            user.set_password(new_password)

        # Mise à jour de la photo
        if 'photo' in request.FILES:
            etudiant.photo = request.FILES['photo']

        user.save()
        etudiant.save()

        return Response({
            "message": "Profil mis à jour",
            "etudiant": {
                'id': etudiant.id,
                'nom': etudiant.nom,
                'prenom': etudiant.prenom,
                'email': user.email,
                'filiere': etudiant.filiere,
                'niveau': etudiant.niveau,
                'photo': etudiant.photo.url if etudiant.photo else None
            }
        }, status=200)



#############################################cote admin########################################
####################Had la vue k ygerer biha l admin les enseignants : crud 

class AdminEnseignantViewSet(viewsets.ModelViewSet):
    queryset = Enseignant.objects.select_related('user').all()
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return EnseignantCreateUpdateSerializer
        return EnseignantListSerializer

    def destroy(self, request, *args, **kwargs):
        enseignant = self.get_object()
        user = enseignant.user
        user.delete()  # Supprime user et enseignant grâce à OneToOne
        return Response({"message": "Enseignant supprimé avec succès"}, status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        enseignant = self.get_object()
        enseignant.user.is_active = True
        enseignant.user.save()
        return Response({'status': 'Compte enseignant activé'})

    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        enseignant = self.get_object()
        enseignant.user.is_active = False
        enseignant.user.save()
        return Response({'status': 'Compte enseignant désactivé'})

################Hna la admin kaychuf les demandes daide okayred 3lihum

class AdminHelpResponseView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request, pk):
        help_request = get_object_or_404(HelpRequest, pk=pk)
        response_text = request.data.get('response')

        if not response_text:
            return Response({'error': 'Réponse vide non autorisée.'}, status=400)

        help_request.response = response_text
        help_request.resolved = True
        help_request.responded_at = timezone.now()
        help_request.save()

        return Response({'status': 'Réponse enregistrée avec succès.'})
    
######################Hadi gi ka ylister biha les demandes li kayjiweh oletat dyalhum
    
class AdminHelpListView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        help_requests = HelpRequest.objects.all().order_by('-created_at')
        serializer = HelpRequestSerializer(help_requests, many=True)
        return Response(serializer.data)
    
##########Hadi 3la 7sab min ygenerer lens qr code twsl notif l admin

class QRNotificationView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    def get(self, request):
        unread = QRNotification.objects.filter(viewed=False).select_related('enseignant', 'session')
        return Response({
            'notifications': [
                {
                    'id': n.id,
                    'enseignant': n.enseignant.email,
                    'session': SessionSerializer(n.session).data,
                    'date': n.created_at
                } for n in unread
            ]
        })

    @action(detail=True, methods=['post'])
    def mark_viewed(self, request, pk=None):
        notification = QRNotification.objects.get(pk=pk)
        notification.viewed = True
        notification.save()
        return Response({'status': 'marked as viewed'})

##########################La vue 3la 7sab y activer les enseignants 
    
class EnseignantViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Enseignant.objects.all()
    serializer_class = EnseignantSerializer
    


###############Les cartes d dashbord li fihum nbr ens nbr etudiant..............

class AdminDashboardStatsView(APIView):
    
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        total_enseignants = Enseignant.objects.count()  
        total_matieres = Matiere.objects.count()
        total_etudiants = Etudiant.objects.count()  

        data = {
            "total_enseignants": total_enseignants,
            "total_matieres": total_matieres,
            "total_etudiants": total_etudiants
        }
        return Response(data)
    
###############Gestion des etudients
class EtudiantViewSet(viewsets.ModelViewSet):
    queryset = Etudiant.objects.all()
    serializer_class = EtudiantSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    filter_backends = [filters.SearchFilter]
    search_fields = ['nom', 'prenom', 'user__email']

    def get_queryset(self):
        queryset = super().get_queryset()
        filiere = self.request.query_params.get('filiere')
        niveau = self.request.query_params.get('niveau')
        
        if filiere:
            queryset = queryset.filter(filiere=filiere)
        if niveau:
            queryset = queryset.filter(niveau=niveau)
            
        return queryset    
    
###############rapport admin/etud
# Nettoyage du texte IA 
def nettoyer_texte_ia(self, texte):
    texte = re.sub(r"<think>.*?</think>", "", texte, flags=re.DOTALL)
    texte = re.sub(r"<\|.*?\|>", "", texte)
    texte = re.sub(r"\n{3,}", "\n\n", texte)  # réduit les sauts de ligne excessifs
    texte = texte.lstrip("\n")  # supprime les sauts de ligne en début de texte
    return texte.strip()


# Générer un résumé IA global
def generer_resume_ia(absences_data, etudiants):
    if not absences_data:
        return "Aucune absence détectée. Impossible de générer une analyse IA pertinente."

    resume = "\n".join([
        f"{a['Nom']} {a['Prénom']} | {a['Module']} | {a['Date']} | Justifiée : {a['Justifiée']}"
        for a in absences_data
    ])

    prompt = f"""
Voici les données d'absences des étudiants (nom, module, date, justification) :

{resume}

Analyse uniquement ces données. Génère un résumé clair pour l’administration avec :
1. Les modules les plus touchés (avec le nombre exact d’absences)
2. Les étudiants les plus absents (avec leur total d’absences)
3. Des recommandations concrètes
4. Une conclusion synthétique

⚠️ Utilise uniquement les noms et matieres présents dans les données. Ne crée pas d’exemples fictifs.

Réponds en français.
"""
    try:
        llm = ChatFireworks(model="accounts/fireworks/models/deepseek-r1", temperature=0.3)
        response = llm.invoke(prompt)
        return nettoyer_texte_ia(response.content)
    except Exception:
        return "Résumé IA indisponible pour le moment."


class ExportAbsencesReport(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def nettoyer_texte_ia(self, texte):
        texte = re.sub(r"<think>.*?</think>", "", texte, flags=re.DOTALL)
        texte = re.sub(r"<\|.*?\|>", "", texte)
        return texte.strip()

    def generer_resume_ia(self, absences_data):
        if not absences_data:
            return "Aucune absence détectée. Impossible de générer une analyse IA pertinente."

        resume = "\n".join([
            f"{a['Nom']} {a['Prénom']} | {a['Module']} | {a['Date']} | Justifiée : {a['Justifiée']}"
            for a in absences_data
        ])

        prompt = f"""
Voici les données d'absences des étudiants (nom, module, date, justification) :

{resume}

Analyse uniquement ces données. Génère un résumé clair pour l’administration avec :
1. Les modules les plus touchés (avec le nombre exact d’absences)
2. Les étudiants les plus absents (avec leur total d’absences)
3. Des recommandations concrètes
4. Une conclusion synthétique

⚠️ Utilise uniquement les noms et matières présents dans les données. Ne crée pas d’exemples fictifs.

Réponds en français.
"""
        try:
            llm = ChatFireworks(model="accounts/fireworks/models/deepseek-r1", temperature=0.3)
            response = llm.invoke(prompt)
            return self.nettoyer_texte_ia(response.content)
        except:
            return "Résumé IA indisponible pour le moment."

    def get(self, request):
        filiere_id = request.query_params.get('filiere')
        niveau_id = request.query_params.get('niveau')
        mois = request.query_params.get('mois')

        if not filiere_id or not niveau_id:
            return Response({"error": "Les paramètres 'filiere' et 'niveau' sont requis."}, status=400)

        try:
            filiere = Filiere.objects.get(id=filiere_id)
            niveau = Niveau.objects.get(id=niveau_id)
        except (Filiere.DoesNotExist, Niveau.DoesNotExist):
            return Response({"error": "Filière ou niveau introuvable."}, status=404)

        sessions = Session.objects.filter(
            matiere__filiere=filiere,
            matiere__niveau=niveau
        ).select_related('matiere', 'salle')

        if mois:
            try:
                annee, mois_num = map(int, mois.split('-'))
                date_debut = datetime(annee, mois_num, 1)
                date_fin = date_debut + relativedelta(months=1)
                sessions = sessions.filter(date_debut__gte=date_debut, date_debut__lt=date_fin)
            except ValueError:
                return Response({"error": "Format de mois invalide. Utilisez YYYY-MM."}, status=400)

        etudiants = Etudiant.objects.filter(
            filiere=filiere.nom,
            niveau=niveau.nom
        ).select_related('user')

        absences_data = []
        for session in sessions:
            presences = Presence.objects.filter(session=session, status='present(e)')
            presents_ids = set(p.etudiant_id for p in presences)

            for etudiant in etudiants:
                if etudiant.id not in presents_ids:
                    absences_data.append({
                        'Nom': etudiant.nom,
                        'Prénom': etudiant.prenom,
                        'Module': session.matiere.nom,
                        'Type': session.type_seance,
                        'Date': session.date_debut.strftime('%d/%m/%Y'),
                        'Heure': f"{session.heure_debut.strftime('%H:%M')} - {session.heure_fin.strftime('%H:%M')}",
                        'Salle': session.salle.nom,
                        'Justifiée': 'Non',
                        'EtudiantID': etudiant.id
                    })

        # Création du fichier Excel
        wb = Workbook()
        wb.remove(wb.active)

        # Feuille Étudiants
        ws_etudiants = wb.create_sheet("Étudiants")
        ws_etudiants.append(["Nom", "Prénom", "Email"])
        for e in etudiants:
            ws_etudiants.append([e.nom, e.prenom, e.user.email if e.user else ""])

        # Feuille Absences
        ws_absences = wb.create_sheet("Absences")
        if absences_data:
            ws_absences.append(list(absences_data[0].keys())[:-1])
            for row in absences_data:
                ws_absences.append([row[k] for k in list(row.keys())[:-1]])
        else:
            ws_absences.append(["Aucune absence trouvée pour les critères sélectionnés."])

        # Résumé IA global
        texte_ia = self.generer_resume_ia(absences_data)
        print("Résumé IA généré :", texte_ia)
        ws_ia = wb.create_sheet("Résumé IA")
        ws_ia["A1"] = "Analyse générée par l'IA"
        ws_ia["A1"].font = Font(bold=True, size=14)
        if texte_ia.strip():
            lignes = [l.strip() for l in texte_ia.splitlines() if l.strip()]
            for i, ligne in enumerate(lignes, start=2):
                ws_ia[f"A{i}"] = ligne
                ws_ia[f"A{i}"].alignment = Alignment(wrap_text=True, vertical="top")
                ws_ia.column_dimensions["A"].width = 100
        else:
                ws_ia["A2"] = "Résumé IA vide ou non généré."
                ws_ia["A2"].alignment = Alignment(wrap_text=True, vertical="top")
                

        # Feuilles individuelles par étudiant
        absences_par_etudiant = defaultdict(lambda: defaultdict(int))
        total_seances_par_module = defaultdict(int)

        for session in sessions:
            total_seances_par_module[session.matiere.nom] += 1

        for a in absences_data:
            absences_par_etudiant[(a['Nom'], a['Prénom'])][a['Module']] += 1

        total_seances = sum(total_seances_par_module.values())

        for etu in etudiants:
            key = (etu.nom, etu.prenom)
            ws_etu = wb.create_sheet(f"{etu.nom}_{etu.prenom}"[:31])
            ws_etu.append(["Module", "Absences non justifiées", "Total séances", "Taux d'absence (%)"])
            total_abs = 0
            for module, total_seance in total_seances_par_module.items():
                nb_abs = absences_par_etudiant[key].get(module, 0)
                taux = (nb_abs / total_seance * 100) if total_seance > 0 else 0
                total_abs += nb_abs
                ws_etu.append([module, nb_abs, total_seance, round(taux, 2)])
            taux_global = (total_abs / total_seances * 100) if total_seances > 0 else 0
            ws_etu.append([])
            ws_etu.append(["", "Taux global d'absence non justifiée (%)", round(taux_global, 2)])

        # Étudiants à risque
        ws_risque = wb.create_sheet("Étudiants à risque")
        ws_risque.append(["Nom", "Prénom", "Absences non justifiées", "Total séances", "Taux global d'absence (%)"])
        for (nom, prenom), modules in absences_par_etudiant.items():
            total_abs = sum(modules.values())
            taux_global = (total_abs / total_seances * 100) if total_seances > 0 else 0
            if taux_global > 40:
                ws_risque.append([nom, prenom, total_abs, total_seances, round(taux_global, 2)])

        # Export final
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"absences_{filiere.nom}_{niveau.nom}_{timezone_now().strftime('%Y%m%d')}.xlsx"
        return HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
#############################################rapport coté admin/ens#######################################
#  Nettoyage du texte IA
def nettoyer_texte_ia(texte):
    texte = re.sub(r"<think>.*?</think>", "", texte, flags=re.DOTALL)
    return texte.strip()

class EnseignantReport(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        periode = request.query_params.get('periode', 'mois')
        today = now().date()

        if periode == 'mois':
            start_date = today.replace(day=1)
            end_date = today
        elif periode == 'semestre':
            start_date = today.replace(month=1 if today.month <= 6 else 7, day=1)
            end_date = today
        else:
            return Response({"error": "Période invalide"}, status=400)

        qrs = QRNotification.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).select_related(
            'enseignant', 'session', 'session__matiere', 'session__salle',
            'session__matiere__filiere', 'session__matiere__niveau'
        )

        enseignants_dict = {}
        for qr in qrs:
            enseignants_dict.setdefault(qr.enseignant, []).append(qr)

        wb = Workbook()
        wb.remove(wb.active)

        for enseignant, qr_list in enseignants_dict.items():
            ws = wb.create_sheet(title=f"{enseignant.nom}_{enseignant.prenom}"[:30])

            titre = f"Rapport des QR Codes générés – Pr. {enseignant.nom} {enseignant.prenom} – Période : {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"
            ws.merge_cells('A1:H1')
            cell = ws['A1']
            cell.value = titre
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal='center')

            headers = ['Date QR', 'Heure début', 'Heure fin', 'Matière', 'Type séance', 'Filière', 'Niveau', 'Salle']
            ws.append(headers)
            for cell in ws[2]:
                cell.font = Font(bold=True)

            for qr in qr_list:
                session = qr.session
                matiere = session.matiere
                ws.append([
                    qr.created_at.date().strftime('%d/%m/%Y'),
                    session.heure_debut.strftime('%H:%M'),
                    session.heure_fin.strftime('%H:%M'),
                    matiere.nom,
                    session.type_seance,
                    matiere.filiere.nom,
                    matiere.niveau.nom,
                    session.salle.nom
                ])

            # Résumé IA par enseignant
            resume = "\n".join([
                f"{qr.created_at.date().strftime('%d/%m/%Y')} – {qr.session.matiere.nom} ({qr.session.type_seance})"
                for qr in qr_list
            ])

            prompt = f"""
Voici les séances de l'enseignant {enseignant.nom} {enseignant.prenom} :

{resume}

Rédige un résumé professionnel sur son activité :
- Fréquence des QR codes
- Modules enseignés
- Régularité des séances
- Recommandations éventuelles

Réponds en français.
"""
            try:
                texte_ia = ChatFireworks(model="accounts/fireworks/models/deepseek-r1").invoke(prompt).content
                texte_ia = nettoyer_texte_ia(texte_ia)
            except:
                texte_ia = "Résumé IA indisponible."

            ws_ia = wb.create_sheet(title=f"Résumé_{enseignant.nom[:20]}")
            ws_ia["A1"] = f"Résumé IA – Pr. {enseignant.nom} {enseignant.prenom}"
            ws_ia["A1"].font = Font(bold=True, size=12)
            ws_ia["A2"] = texte_ia
            ws_ia["A2"].alignment = Alignment(wrap_text=True)

        # Analyse IA globale sur les QR codes
        try:
            qr_counts = {
                f"{e.nom} {e.prenom}": len(qrs)
                for e in enseignants_dict.keys()
            }
            resume_qr = "\n".join([f"{nom} : {count} QR générés" for nom, count in qr_counts.items()])

            prompt_global = f"""
Voici le nombre de QR codes générés par chaque enseignant :

{resume_qr}

Analyse ces données :
- Qui génère peu de QR ?
- Y a-t-il des anomalies ?
- Que recommanderiez-vous à l’administration ?

Réponds en français.
"""
            interpretation = ChatFireworks(model="accounts/fireworks/models/deepseek-r1").invoke(prompt_global).content
            interpretation = nettoyer_texte_ia(interpretation)
        except:
            interpretation = "Analyse IA indisponible."

        ws_global = wb.create_sheet("Analyse QR IA")
        ws_global["A1"] = "Analyse IA sur la génération de QR codes"
        ws_global["A1"].font = Font(bold=True)
        ws_global["A2"] = interpretation
        ws_global["A2"].alignment = Alignment(wrap_text=True)

        # Alerte IA pour enseignants inactifs
        enseignants_actifs = set(enseignants_dict.keys())
        enseignants_tous = Enseignant.objects.all()
        enseignants_inactifs = [e for e in enseignants_tous if e not in enseignants_actifs]

        if enseignants_inactifs:
            noms_inactifs = "\n".join([f"{e.nom} {e.prenom}" for e in enseignants_inactifs])

            prompt_inactifs = f"""
Voici la liste des enseignants qui n'ont généré aucun QR code pendant la période :

{noms_inactifs}

Rédige une alerte professionnelle à destination de l’administration :
- Pourquoi est-ce problématique ?
- Que recommander ?
- Ton neutre, mais ferme.

Réponds en français.
"""
            try:
                alerte = ChatFireworks(model="accounts/fireworks/models/deepseek-r1").invoke(prompt_inactifs).content
                alerte = nettoyer_texte_ia(alerte)
            except:
                alerte = "Alerte IA indisponible."

            ws_alerte = wb.create_sheet("Alerte enseignants inactifs")
            ws_alerte["A1"] = "Alerte IA – Enseignants sans QR"
            ws_alerte["A1"].font = Font(bold=True)
            ws_alerte["A2"] = alerte
            ws_alerte["A2"].alignment = Alignment(wrap_text=True)

        #  Export
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"rapport_qr_professeurs_{periode}_{now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)
        return response
#==============================================Autre views============================================================
    
class FiliereViewSet(viewsets.ModelViewSet):
    queryset = Filiere.objects.all()
    serializer_class = FiliereSerializer
    permission_classes = [AllowAny] 

class NiveauViewSet(viewsets.ModelViewSet):
    queryset = Niveau.objects.all()
    serializer_class = NiveauSerializer
    permission_classes = [AllowAny]
    
###############Hadi 3la 7sab fil w niveau okadiri ta ens 
class MatiereViewSet(viewsets.ModelViewSet):
    permission_classes = [AllowAny]
    queryset = Matiere.objects.all()
    serializer_class = MatiereSerializer
    def get_queryset(self):
        user = self.request.user
        queryset = Matiere.objects.all()

        # Si l'utilisateur est un enseignant, filtrer ses matières
        if hasattr(user, 'enseignant'):
            queryset = queryset.filter(enseignant=user.enseignant)

        # Filtres supplémentaires
        filiere_id = self.request.query_params.get('filiere')
        niveau_id = self.request.query_params.get('niveau')
        if filiere_id:
            queryset = queryset.filter(filiere__id=filiere_id)
        if niveau_id:
            queryset = queryset.filter(niveau__id=niveau_id)

        return queryset




class SalleViewSet(viewsets.ModelViewSet):
    queryset = Salle.objects.all()
    serializer_class = SalleSerializer

 #############logout li nsit nzido fpartie dyali    
@api_view(['POST'])
def logout_view(request):
    try:
        refresh_token = request.data.get("refresh")
        if not refresh_token:
            return Response({"error": "Token de rafraîchissement requis"}, status=400)

        token = RefreshToken(refresh_token)
        token.blacklist()

        return Response({"message": "Déconnexion réussie"}, status=200)
    except Exception as e:
        return Response({"error": str(e)}, status=500)
    



